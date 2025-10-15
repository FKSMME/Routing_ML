"""
Data Mapping Service - 라우팅 그룹 필드 ↔ DB 컬럼 매핑 관리.

관리자가 라우팅 그룹 데이터를 DB 컬럼 구조로 변환하는 매핑 규칙을 정의하고 관리합니다.
"""

from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
from uuid import uuid4

from backend.api.schemas import (
    DataMappingApplyRequest,
    DataMappingApplyResponse,
    DataMappingProfile,
    DataMappingProfileCreate,
    DataMappingProfileListResponse,
    DataMappingProfileUpdate,
    DataMappingRule,
)

logger = logging.getLogger(__name__)


class DataMappingService:
    """데이터 매핑 프로파일 관리 서비스."""

    def __init__(self, storage_path: Optional[Path] = None):
        """
        Args:
            storage_path: 매핑 프로파일 저장 경로 (기본값: config/data_mappings/)
        """
        self.storage_path = storage_path or Path("config/data_mappings")
        self.storage_path.mkdir(parents=True, exist_ok=True)
        self._profiles: Dict[str, DataMappingProfile] = {}
        self._load_profiles()

    def _load_profiles(self) -> None:
        """저장된 모든 프로파일을 로드합니다."""
        try:
            for file_path in self.storage_path.glob("*.json"):
                try:
                    with open(file_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                        profile = DataMappingProfile(**data)
                        self._profiles[profile.id] = profile
                except Exception as e:
                    logger.warning(f"Failed to load profile from {file_path}: {e}")
            logger.info(f"Loaded {len(self._profiles)} data mapping profiles")
        except Exception as e:
            logger.error(f"Failed to load profiles: {e}")

    def _save_profile(self, profile: DataMappingProfile) -> None:
        """프로파일을 파일로 저장합니다."""
        file_path = self.storage_path / f"{profile.id}.json"
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(profile.model_dump(), f, indent=2, ensure_ascii=False, default=str)
        logger.info(f"Saved profile {profile.id} to {file_path}")

    def _delete_profile_file(self, profile_id: str) -> None:
        """프로파일 파일을 삭제합니다."""
        file_path = self.storage_path / f"{profile_id}.json"
        if file_path.exists():
            file_path.unlink()
            logger.info(f"Deleted profile file {file_path}")

    def list_profiles(self) -> DataMappingProfileListResponse:
        """모든 프로파일 목록을 반환합니다."""
        profiles = list(self._profiles.values())
        # 최신 수정 순으로 정렬
        profiles.sort(
            key=lambda p: p.updated_at or p.created_at or datetime.min,
            reverse=True,
        )
        return DataMappingProfileListResponse(
            profiles=profiles,
            total=len(profiles),
        )

    def get_profile(self, profile_id: str) -> Optional[DataMappingProfile]:
        """특정 프로파일을 조회합니다."""
        return self._profiles.get(profile_id)

    def create_profile(
        self,
        payload: DataMappingProfileCreate,
        created_by: str,
    ) -> DataMappingProfile:
        """새 프로파일을 생성합니다."""
        now = datetime.utcnow()
        profile_id = str(uuid4())

        profile = DataMappingProfile(
            id=profile_id,
            name=payload.name,
            description=payload.description,
            mappings=payload.mappings,
            created_by=created_by,
            created_at=now,
            updated_at=now,
            is_active=True,
        )

        self._profiles[profile_id] = profile
        self._save_profile(profile)

        logger.info(
            f"Created data mapping profile: {profile_id} by {created_by}",
            extra={"profile_id": profile_id, "created_by": created_by},
        )
        return profile

    def update_profile(
        self,
        profile_id: str,
        payload: DataMappingProfileUpdate,
    ) -> Optional[DataMappingProfile]:
        """기존 프로파일을 수정합니다."""
        profile = self._profiles.get(profile_id)
        if not profile:
            return None

        now = datetime.utcnow()

        # 수정할 필드만 업데이트
        if payload.name is not None:
            profile.name = payload.name
        if payload.description is not None:
            profile.description = payload.description
        if payload.mappings is not None:
            profile.mappings = payload.mappings
        if payload.is_active is not None:
            profile.is_active = payload.is_active

        profile.updated_at = now

        self._save_profile(profile)

        logger.info(
            f"Updated data mapping profile: {profile_id}",
            extra={"profile_id": profile_id},
        )
        return profile

    def delete_profile(self, profile_id: str) -> bool:
        """프로파일을 삭제합니다."""
        if profile_id not in self._profiles:
            return False

        del self._profiles[profile_id]
        self._delete_profile_file(profile_id)

        logger.info(
            f"Deleted data mapping profile: {profile_id}",
            extra={"profile_id": profile_id},
        )
        return True

    def apply_mapping(
        self,
        request: DataMappingApplyRequest,
        routing_group_data: List[Dict[str, Any]],
    ) -> DataMappingApplyResponse:
        """
        라우팅 그룹 데이터에 매핑 프로파일을 적용하여 변환합니다.

        Args:
            request: 매핑 적용 요청
            routing_group_data: 라우팅 그룹 데이터 (공정 단계 목록)

        Returns:
            DataMappingApplyResponse: 변환된 데이터 및 미리보기
        """
        profile = self.get_profile(request.profile_id)
        if not profile:
            raise ValueError(f"Profile not found: {request.profile_id}")

        # 매핑 규칙 적용
        transformed_rows: List[Dict[str, Any]] = []
        columns: List[str] = []

        # 컬럼명 추출 (매핑 규칙의 display_name 또는 db_column)
        for rule in profile.mappings:
            col_name = rule.display_name or rule.db_column
            if col_name not in columns:
                columns.append(col_name)

        # 각 행 변환
        for row_data in routing_group_data:
            transformed_row: Dict[str, Any] = {}

            for rule in profile.mappings:
                col_name = rule.display_name or rule.db_column

                # source_type에 따라 값 추출
                value = self._extract_value_by_source_type(rule, row_data)

                # 데이터 타입 변환
                if value is not None:
                    value = self._convert_value(value, rule.data_type)

                transformed_row[col_name] = value

            transformed_rows.append(transformed_row)

        # 미리보기는 최대 10행만
        preview_rows = transformed_rows[:10]

        csv_path: Optional[str] = None
        if not request.preview_only:
            # 파일 생성 (CSV 또는 XLSX)
            export_format = getattr(request, 'export_format', 'csv')

            if export_format == 'xlsx':
                csv_path = self._export_to_xlsx(
                    columns=columns,
                    rows=transformed_rows,
                    routing_group_id=request.routing_group_id,
                    profile_id=request.profile_id,
                )
            else:
                csv_path = self._export_to_csv(
                    columns=columns,
                    rows=transformed_rows,
                    routing_group_id=request.routing_group_id,
                    profile_id=request.profile_id,
                )

        return DataMappingApplyResponse(
            profile_id=request.profile_id,
            routing_group_id=request.routing_group_id,
            columns=columns,
            preview_rows=preview_rows,
            total_rows=len(transformed_rows),
            csv_path=csv_path,
            message=(
                f"미리보기 생성 완료 ({len(transformed_rows)}행)"
                if request.preview_only
                else f"{'XLSX' if export_format == 'xlsx' else 'CSV'} 파일 생성 완료: {csv_path}"
            ),
        )

    def _extract_value_by_source_type(
        self,
        rule: DataMappingRule,
        row_data: Dict[str, Any],
    ) -> Any:
        """
        데이터 소스 타입에 따라 값을 추출합니다.

        Args:
            rule: 매핑 규칙
            row_data: 라우팅 그룹 데이터 행

        Returns:
            추출된 값
        """
        source_type = getattr(rule, 'source_type', 'ml_prediction')

        if source_type == 'ml_prediction':
            # ML 예측 결과에서 가져옴 (timeline 데이터)
            return row_data.get(rule.routing_field, rule.default_value)

        elif source_type == 'admin_input':
            # 관리자가 설정한 고정 값 사용
            return rule.default_value

        elif source_type == 'constant':
            # 모든 행에 동일한 고정 값
            return rule.default_value

        elif source_type == 'external_source':
            # 외부 소스에서 가져옴 (공정그룹 관리 등)
            source_config = getattr(rule, 'source_config', None)
            if source_config and isinstance(source_config, dict):
                source_cfg_type = source_config.get('type')

                if source_cfg_type == 'process_group':
                    # 프로세스 그룹에서 값 가져오기
                    value = self._get_value_from_process_group(
                        source_config=source_config,
                        row_data=row_data,
                    )
                    if value is not None:
                        return value
                else:
                    # 다른 external_source 타입은 row_data에서 먼저 찾음
                    external_field = source_config.get('field', rule.routing_field)
                    value = row_data.get(external_field)
                    if value is not None:
                        return value

            # 외부 소스 연동이 안 되어 있으면 기본값 사용
            return rule.default_value

        # 알 수 없는 타입은 ml_prediction과 동일하게 처리
        return row_data.get(rule.routing_field, rule.default_value)

    def _get_value_from_process_group(
        self,
        source_config: Dict[str, Any],
        row_data: Dict[str, Any],
    ) -> Optional[Any]:
        """
        프로세스 그룹에서 값을 가져옵니다.

        Args:
            source_config: 소스 설정 {'type': 'process_group', 'group_id': '...', 'field': '...'}
            row_data: 현재 행 데이터

        Returns:
            프로세스 그룹에서 가져온 값 또는 None
        """
        try:
            from backend.models.process_groups import ProcessGroup, session_scope

            group_id = source_config.get('group_id')
            field_key = source_config.get('field')

            if not group_id or not field_key:
                logger.warning(
                    f"Invalid process_group source_config: {source_config}"
                )
                return None

            # 프로세스 그룹 조회
            with session_scope() as session:
                group = (
                    session.query(ProcessGroup)
                    .filter(
                        ProcessGroup.id == group_id,
                        ProcessGroup.deleted_at.is_(None),
                        ProcessGroup.is_active == True,
                    )
                    .first()
                )

                if not group:
                    logger.warning(f"Process group not found: {group_id}")
                    return None

                # fixed_values에서 값 찾기
                fixed_values = group.fixed_values or {}
                if field_key in fixed_values:
                    return fixed_values[field_key]

                # 컬럼 정의에서 기본값 찾기
                default_columns = group.default_columns or []
                for col in default_columns:
                    if col.get('key') == field_key:
                        # 컬럼이 존재하지만 fixed_values에 없으면 None 반환
                        return None

                logger.debug(
                    f"Field '{field_key}' not found in process group {group_id}"
                )
                return None

        except Exception as e:
            logger.error(
                f"Failed to get value from process group: {e}",
                exc_info=True,
            )
            return None

    def _convert_value(self, value: Any, data_type: str) -> Any:
        """값을 지정된 데이터 타입으로 변환합니다."""
        try:
            if data_type == "number":
                return float(value) if value not in (None, "", "nan") else None
            elif data_type == "boolean":
                if isinstance(value, bool):
                    return value
                if isinstance(value, str):
                    return value.lower() in ("true", "yes", "1", "y")
                return bool(value)
            elif data_type == "date":
                # 날짜 문자열을 그대로 반환 (추후 포맷 변환 가능)
                return str(value) if value not in (None, "", "nan") else None
            else:  # string
                return str(value) if value not in (None, "", "nan") else ""
        except Exception as e:
            logger.warning(f"Failed to convert value {value} to {data_type}: {e}")
            return value

    def _export_to_csv(
        self,
        columns: List[str],
        rows: List[Dict[str, Any]],
        routing_group_id: str,
        profile_id: str,
    ) -> str:
        """데이터를 CSV 파일로 내보냅니다."""
        import csv

        # 출력 디렉토리 생성
        output_dir = Path("output/comprehensive_routing")
        output_dir.mkdir(parents=True, exist_ok=True)

        # 파일명 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"routing_{routing_group_id}_{timestamp}.csv"
        file_path = output_dir / filename

        # CSV 작성
        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            writer.writerows(rows)

        logger.info(
            f"Exported CSV: {file_path} ({len(rows)} rows)",
            extra={
                "routing_group_id": routing_group_id,
                "profile_id": profile_id,
                "rows": len(rows),
            },
        )
        return str(file_path)

    def _export_to_xlsx(
        self,
        columns: List[str],
        rows: List[Dict[str, Any]],
        routing_group_id: str,
        profile_id: str,
    ) -> str:
        """
        데이터를 XLSX 파일로 내보냅니다 (4개 시트 분할).

        시트 구조:
        - 병합1: 전체 181개 컬럼 (MASTER + DETAIL + CHILD + RES)
        - P_BOP_PROC_MASTER: 컬럼 1-28
        - P_BOP_PROC_DETAIL: 컬럼 29-121 (93개)
        - P_BOP_PROC_CHILD: 컬럼 122-142 (21개)
        - P_BOP_PROC_RES: 컬럼 143-181 (39개)
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        # 출력 디렉토리 생성
        output_dir = Path("output/comprehensive_routing")
        output_dir.mkdir(parents=True, exist_ok=True)

        # 파일명 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"routing_{routing_group_id}_{timestamp}.xlsx"
        file_path = output_dir / filename

        # Workbook 생성
        wb = Workbook()

        # 시트 분할 정의 (컬럼 인덱스 기준)
        sheet_config = {
            "병합1": {"start": 0, "end": len(columns), "columns": columns},
            "P_BOP_PROC_MASTER": {"start": 0, "end": 28, "columns": columns[0:28]},
            "P_BOP_PROC_DETAIL": {"start": 28, "end": 121, "columns": columns[28:121]},
            "P_BOP_PROC_CHILD": {"start": 121, "end": 142, "columns": columns[121:142]},
            "P_BOP_PROC_RES": {"start": 142, "end": 181, "columns": columns[142:181]},
        }

        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 각 시트 생성
        for sheet_idx, (sheet_name, config) in enumerate(sheet_config.items()):
            if sheet_idx == 0:
                # 첫 번째 시트는 기본 시트 이름 변경
                ws = wb.active
                ws.title = sheet_name
            else:
                # 새 시트 추가
                ws = wb.create_sheet(title=sheet_name)

            sheet_columns = config["columns"]
            if len(sheet_columns) == 0:
                # 빈 시트는 건너뛰기
                continue

            # 헤더 작성
            for col_idx, col_name in enumerate(sheet_columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 데이터 작성
            for row_idx, row_data in enumerate(rows, start=2):
                for col_idx, col_name in enumerate(sheet_columns, start=1):
                    value = row_data.get(col_name, "")
                    # None이나 "nan"을 빈 문자열로 변환
                    if value is None or value == "nan" or value == "NULL":
                        value = ""
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # 컬럼 너비 자동 조정
            for col_idx, col_name in enumerate(sheet_columns, start=1):
                # 컬럼명 길이 기준으로 너비 설정 (최소 10, 최대 50)
                col_width = min(max(len(str(col_name)) + 2, 10), 50)
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col_width

        # 파일 저장
        wb.save(file_path)

        logger.info(
            f"Exported XLSX: {file_path} ({len(rows)} rows, {len(sheet_config)} sheets)",
            extra={
                "routing_group_id": routing_group_id,
                "profile_id": profile_id,
                "rows": len(rows),
                "sheets": len(sheet_config),
            },
        )
        return str(file_path)


# 싱글톤 인스턴스
_data_mapping_service: Optional[DataMappingService] = None


def get_data_mapping_service() -> DataMappingService:
    """DataMappingService 싱글톤 인스턴스를 반환합니다."""
    global _data_mapping_service
    if _data_mapping_service is None:
        _data_mapping_service = DataMappingService()
    return _data_mapping_service
