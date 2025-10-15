import openpyxl
import json

wb = openpyxl.load_workbook(r'c:\Users\syyun\Documents\GitHub\Routing_ML_251014\FKSM 라우팅 데이터구조.xlsx', data_only=True)

# 두 번째 시트 (병합1)
ws = wb.worksheets[1]  # 인덱스로 접근
print(f'Sheet name: {ws.title}')
print(f'Size: {ws.max_row} rows x {ws.max_column} cols')
print(f'Merged cells: {len(list(ws.merged_cells))}\n')

# 헤더 행 찾기
print('=== First 10 rows (first 30 columns) ===')
for i in range(1, 11):
    row_values = []
    for j in range(1, 31):
        cell = ws.cell(row=i, column=j)
        val = str(cell.value) if cell.value is not None else ''
        row_values.append(val[:30])  # 30자까지만
    if any(row_values):  # 빈 행이 아니면 출력
        print(f'R{i:2d}: {" | ".join(row_values[:15])}')

print('\n=== All Column Headers ===')
header_row = 1  # 헤더가 1행이라고 가정
headers = []
for col_idx in range(1, ws.max_column + 1):
    cell_value = ws.cell(row=header_row, column=col_idx).value
    header_name = str(cell_value) if cell_value is not None else f'Col_{col_idx}'
    headers.append(header_name)
    if col_idx <= 50:  # 처음 50개만 출력
        print(f'{col_idx:3d}. {header_name}')

# JSON 저장
result = {
    'sheet_name': ws.title,
    'total_rows': ws.max_row,
    'total_columns': ws.max_column,
    'merged_cells_count': len(list(ws.merged_cells)),
    'headers': headers
}

with open('merged_sheet_columns.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, indent=2, ensure_ascii=False)

print(f'\n\nTotal columns: {len(headers)}')
print('Saved to merged_sheet_columns.json')
