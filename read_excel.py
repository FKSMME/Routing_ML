import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\syyun\Documents\GitHub\Routing_ML_251014\FKSM 라우팅 데이터구조.xlsx', data_only=True)
print('Sheet names:', wb.sheetnames)
ws = wb.active
print('\nActive sheet:', ws.title)
print('\nMerged cells:')
for merged_cell in list(ws.merged_cells)[:20]:
    print(f'  {merged_cell}')

print('\n\nFirst 30 rows (first 15 columns):')
for i, row in enumerate(ws.iter_rows(max_row=30, values_only=True)):
    cells = [str(c) if c is not None else '' for c in row[:15]]
    print(f'Row {i+1:2d}: {cells}')
