import openpyxl
import json

wb = openpyxl.load_workbook(r'c:\Users\syyun\Documents\GitHub\Routing_ML_251014\FKSM 라우팅 데이터구조.xlsx', data_only=True)

# P_BOP_PROC_DETAIL 시트가 가장 많은 컬럼 (93개)을 가지고 있음
sheets_to_analyze = ['P_BOP_PROC_DETAIL', 'P_BOP_PROC_RES', 'P_BOP_PROC_MASTER', 'P_BOP_PROC_CHILD']

result = {}
for sheet_name in sheets_to_analyze:
    ws = wb[sheet_name]

    # 헤더 행 (첫 번째 행)
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    columns = [str(c) if c is not None else f'Col_{i}' for i, c in enumerate(header_row)]

    # 샘플 데이터 (2번째 행)
    sample_row = None
    for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
        sample_row = [str(c) if c is not None else '' for c in row]
        break

    result[sheet_name] = {
        'total_columns': len(columns),
        'columns': columns,
        'sample_data': sample_row
    }

# JSON 파일로 저장
with open('fksm_columns.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, indent=2, ensure_ascii=False)

print('Saved to fksm_columns.json\n')

# P_BOP_PROC_DETAIL의 모든 컬럼 출력
print('=== P_BOP_PROC_DETAIL Columns (93개) ===')
for i, col in enumerate(result['P_BOP_PROC_DETAIL']['columns'], 1):
    print(f'{i:3d}. {col}')
