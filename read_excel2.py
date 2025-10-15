import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\syyun\Documents\GitHub\Routing_ML_251014\FKSM 라우팅 데이터구조.xlsx', data_only=True)

print('=== All Sheets ===')
for idx, sheet_name in enumerate(wb.sheetnames):
    ws = wb[sheet_name]
    print(f'\n[{idx+1}] {sheet_name}')
    print(f'    Size: {ws.max_row} rows x {ws.max_column} cols')
    print(f'    Merged cells: {len(list(ws.merged_cells))}')

    # 첫 5행만 미리보기
    print(f'    Preview (first 5 rows, first 10 cols):')
    for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
        cells = [str(c)[:20] if c is not None else '' for c in row[:10]]
        if any(cells):  # 빈 행이 아니면 출력
            print(f'      R{i+1}: {cells}')

# P_BOP_PROC_RES 시트 상세 분석
print('\n\n=== P_BOP_PROC_RES Sheet Detail ===')
ws = wb['P_BOP_PROC_RES']
print(f'Merged cells:')
for m in list(ws.merged_cells)[:20]:
    print(f'  {m}')

print(f'\nFirst 15 rows:')
for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True)):
    cells = [str(c)[:30] if c is not None else '' for c in row[:15]]
    print(f'Row {i+1:2d}: {" | ".join(cells)}')
